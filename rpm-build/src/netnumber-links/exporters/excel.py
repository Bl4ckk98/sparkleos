from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette colori
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
NORMAL_ROW_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
CELL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def write_sheet(ws, rows: list[dict], columns: list[str]):
    """Scrive i dati su un foglio Excel con formattazione."""
    # Intestazione
    ws.append(columns)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Dati
    for row_idx, row_data in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else NORMAL_ROW_FILL
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, "N/A"))
            cell.fill = fill
            cell.font = CELL_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = THIN_BORDER

    # Adatta la larghezza delle colonne
    col_widths = [len(c) for c in columns]
    for row_data in rows:
        for col_idx, col_name in enumerate(columns):
            val = str(row_data.get(col_name, ""))
            if len(val) > col_widths[col_idx]:
                col_widths[col_idx] = len(val)

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 60)

    # Altezza riga header
    ws.row_dimensions[1].height = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    # Filtro automatico
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
